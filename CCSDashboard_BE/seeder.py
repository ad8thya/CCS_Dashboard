"""
CCS Data Seeder
Reads the three Excel files and inserts into PostgreSQL.
Run: python seed_ccs.py
"""

import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from datetime import datetime, date

# ─── CONFIG ──────────────────────────────────────────────────────────────────
DB = {
    "host":     "localhost",
    "port":     5432,
    "dbname":   "CCSDashboard",
    "user":     "postgres",
    "password": "password",
}

BATCH_FILE  = r"C:\Users\adisi\Downloads\Batch_Report_Summary_20260601_1216.xlsx"
CERT_FILE   = r"C:\Users\adisi\Downloads\Active_Certificates_2026-06-01_to_2026-06-30_20260601_1219.xlsx"
COMP_FILE   = r"C:\Users\adisi\Downloads\Competency_Register.xlsx"

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def xl_serial_to_date(value) -> date:
    """Convert Excel serial number OR already-a-date to a Python date."""
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    if isinstance(value, (int, float)):
        # Excel epoch: Dec 30 1899  (accounts for Lotus 1-2-3 leap-year bug)
        return (datetime(1899, 12, 30) + __import__('datetime').timedelta(days=int(value))).date()
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    raise ValueError(f"Cannot parse date: {value!r}")


def rows(sheet, skip_header_rows=3):
    """Yield non-empty data rows, skipping the first N header rows."""
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i < skip_header_rows:
            continue
        if all(v is None for v in row):
            continue
        yield row


# ─── PARSE ───────────────────────────────────────────────────────────────────

def parse_batches():
    wb = load_workbook(BATCH_FILE, read_only=True, data_only=True)
    ws = wb.active
    batches = []
    for row in rows(ws, skip_header_rows=3):
        # columns: S.No | Batch | Training Type | Department | Category | From | To | Trainees
        if row[0] is None:
            continue
        batches.append({
            "batch_code":   str(row[1]).strip(),
            "course_name":  str(row[2]).strip(),   # Training Type as course name
            "department":   str(row[3]).strip(),
            "category":     str(row[4]).strip(),
            "start_date":   xl_serial_to_date(row[5]),
            "end_date":     xl_serial_to_date(row[6]),
            "trainer":      "",                     # not in source file
            "venue":        "",
        })
    wb.close()
    return batches


def parse_employees_and_certs():
    """
    Competency Register has the richest employee data.
    Active Certificates has the area (competency area) per cert.
    We merge on CertificateNo.
    """
    # --- Active Certificates (area field) ---
    wb_cert = load_workbook(CERT_FILE, read_only=True, data_only=True)
    ws_cert = wb_cert.active
    # columns: S.No | Staff ID | Employee | Department | Contractor | Area | Cert No | Valid From | Valid Until
    area_map = {}   # cert_no -> area
    dept_map = {}   # staff_id -> department
    contractor_map = {}  # staff_id -> contractor
    for row in rows(ws_cert, skip_header_rows=3):
        if row[0] is None:
            continue
        cert_no  = str(row[6]).strip()
        area_map[cert_no] = str(row[5]).strip()
        staff_id = str(row[1]).strip()
        dept_map[staff_id]       = str(row[3]).strip()
        contractor_map[staff_id] = str(row[4]).strip()
    wb_cert.close()

    # --- Competency Register (master employee + cert data) ---
    wb_comp = load_workbook(COMP_FILE, read_only=True, data_only=True)
    ws_comp = wb_comp.active
    # columns: S.No | Name | Designation | Staff ID | DOB | Cert No | Company | Date of Issue | Date of Validity | Remarks
    employees = {}   # staff_id -> dict
    certs     = []

    for row in rows(ws_comp, skip_header_rows=3):
        if row[0] is None:
            continue

        name        = str(row[1]).strip()
        designation = str(row[2]).strip()
        staff_id    = str(row[3]).strip()
        dob_raw     = row[4]
        cert_no     = str(row[5]).strip()
        company     = str(row[6]).strip()
        issue_raw   = row[7]
        valid_raw   = row[8]

        dept = dept_map.get(staff_id, "")

        if staff_id not in employees:
            employees[staff_id] = {
                "employee_code": staff_id,
                "name":          name,
                "designation":   designation,
                "department":    dept,
                "contractor":    contractor_map.get(staff_id, company),
            }

        certs.append({
            "certificate_number": cert_no,
            "employee_code":      staff_id,
            "issue_date":         xl_serial_to_date(issue_raw),
            "expiry_date":        xl_serial_to_date(valid_raw),
            "competency_area":    area_map.get(cert_no, ""),
            "status":             "Active",
        })

    wb_comp.close()
    return list(employees.values()), certs


# ─── INSERT ──────────────────────────────────────────────────────────────────

def seed():
    batches           = parse_batches()
    employees, certs  = parse_employees_and_certs()

    conn = psycopg2.connect(**DB)
    cur  = conn.cursor()

    # ── Employees ────────────────────────────────────────────────────────────
    print(f"Inserting {len(employees)} employees...")
    execute_values(cur, """
        INSERT INTO "Employees" ("EmployeeCode", "Name", "Designation", "Department")
        VALUES %s
        ON CONFLICT DO NOTHING
    """, [(e["employee_code"], e["name"], e["designation"], e["department"])
          for e in employees])

    # ── Batches ──────────────────────────────────────────────────────────────
    print(f"Inserting {len(batches)} batches...")
    execute_values(cur, """
        INSERT INTO "Batches" ("BatchCode", "CourseName", "StartDate", "EndDate", "Trainer", "Venue")
        VALUES %s
        ON CONFLICT DO NOTHING
    """, [(b["batch_code"], b["course_name"], b["start_date"],
           b["end_date"],   b["trainer"],     b["venue"])
          for b in batches])

    # ── Build lookup maps (after insert) ─────────────────────────────────────
    cur.execute("""SELECT "EmployeeCode", "Id" FROM "Employees" """)
    emp_id_map = {row[0]: row[1] for row in cur.fetchall()}

    # Map cert → batch via the Batch Report's batch_code list
    # The Active Certs sheet doesn't have batch info, so we derive it:
    # cert prefix encodes department/type. Use first matching batch by dept.
    cur.execute("""SELECT "BatchCode", "Id", "CourseName" FROM "Batches" """)
    batch_rows = cur.fetchall()
    batch_id_map = {r[0]: r[1] for r in batch_rows}

    # Dept → BatchCode heuristic mapping from the Batch Report data
    dept_to_batch = {}
    for b in batches:
        dept_to_batch.setdefault(b["department"], b["batch_code"])

    # ── Certificates ─────────────────────────────────────────────────────────
    print(f"Inserting {len(certs)} certificates...")
    cert_rows = []
    for c in certs:
        emp_id = emp_id_map.get(c["employee_code"])
        if emp_id is None:
            print(f"  WARN: no employee found for code {c['employee_code']!r}, skipping cert {c['certificate_number']}")
            continue

        # Derive batch from competency area / employee dept
        emp = next((e for e in employees if e["employee_code"] == c["employee_code"]), None)
        dept = emp["department"] if emp else ""
        batch_code = dept_to_batch.get(dept, batches[0]["batch_code"] if batches else None)
        batch_id   = batch_id_map.get(batch_code) if batch_code else None

        if batch_id is None:
            print(f"  WARN: no batch found for dept {dept!r}, skipping cert {c['certificate_number']}")
            continue

        cert_rows.append((
            c["certificate_number"],
            emp_id,
            batch_id,
            c["issue_date"],
            c["expiry_date"],
            c["status"],
            c["competency_area"],
        ))

    execute_values(cur, """
        INSERT INTO "Certificates"
            ("CertificateNumber", "EmployeeId", "BatchId",
             "IssueDate", "ExpiryDate", "Status", "CompetencyArea")
        VALUES %s
        ON CONFLICT DO NOTHING
    """, cert_rows)

    conn.commit()
    cur.close()
    conn.close()
    print("Done. All data seeded successfully.")


if __name__ == "__main__":
    seed()